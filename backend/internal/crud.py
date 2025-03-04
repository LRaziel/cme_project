from sqlalchemy.orm import Session
from internal.models.db_models import User, Material, Tracking, Failure
from internal.schemas import UserCreate, MaterialCreate, TrackingCreate, FailureCreate
import hashlib
from fpdf import FPDF
import openpyxl

# CRUD para Usuários

# Obtém todos os usuários
def get_users(db: Session):
    return db.query(User).all()

# Obtém usuários por função
def get_users_by_role(db: Session, role: str):
    return db.query(User).filter(User.role == role).all()

# Cria um novo usuário
def create_new_user(db: Session, user: UserCreate):
    db_user = User(name=user.name, email=user.email, password=user.password, role=user.role)
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

# CRUD para Materiais

# Gera um serial único para o material
def generate_serial(name: str) -> str:
    return hashlib.md5(name.encode()).hexdigest()

# Obtém todos os materiais
def get_materials(db: Session):
    return db.query(Material).all()

# Obtém materiais por etapa
def get_materials_by_stage(db: Session, stage: str):
    return db.query(Material).join(Tracking).filter(Tracking.stage == stage).all()

# Obtém materiais por status
def get_materials_by_status(db: Session, status: str):
    return db.query(Material).join(Tracking).filter(Tracking.status == status).all()

# Cria um novo material
def create_new_material(db: Session, material: MaterialCreate):
    serial = generate_serial(material.name)
    db_material = Material(name=material.name, type=material.type, expiration_date=material.expiration_date, serial=serial)
    db.add(db_material)
    db.commit()
    db.refresh(db_material)
    return db_material

# Obtém materiais com rastreamento
def get_materials_with_tracking(db: Session):
    materials = db.query(Material).all()
    result = []
    for material in materials:
      tracking_records = get_tracking_by_serial(db, material.serial)
      for tracking in tracking_records:
        if tracking.material_id == material.id:
          result.append({
            "id": material.id,
            "name": material.name,
            "type": material.type,
            "expiration_date": material.expiration_date,
            "serial": material.serial,
            "stage": tracking.stage,
            "status": tracking.status,
          })
        else:
          result.append({
            "id": material.id,
            "name": material.name,
            "type": material.type,
            "expiration_date": material.expiration_date,
            "serial": material.serial,
            "stage": None,
            "status": None,
          })
    return result

# CRUD para Rastreamento

# Obtém todos os rastreamentos
def get_tracking(db: Session):
    return db.query(Tracking).all()

# Obtém rastreamentos por serial
def get_tracking_by_serial(db: Session, serial: str):
    return db.query(Tracking).join(Material).filter(Material.serial == serial).all()

# Obtém falhas por serial
def get_failures_by_serial(db: Session, serial: str):
    return db.query(Failure).join(Tracking).join(Material).filter(Material.serial == serial).all()

# Cria um novo rastreamento
def create_new_tracking(db: Session, tracking: TrackingCreate):
    db_tracking = Tracking(material_id=tracking.material_id, stage=tracking.stage, status=tracking.status)
    db.add(db_tracking)
    db.commit()
    db.refresh(db_tracking)
    return db_tracking

# CRUD para Falhas

# Obtém todas as falhas
def get_failures(db: Session):
    return db.query(Failure).all()

# Cria uma nova falha
def create_failure(db: Session, failure: FailureCreate):
    db_failure = Failure(tracking_id=failure.tracking_id, reason=failure.reason)
    db.add(db_failure)
    db.commit()
    db.refresh(db_failure)
    return db_failure

# Funções para gerar relatórios

# Gera um relatório PDF de rastreamento
def generate_pdf_report(db: Session):
    tracking_data = get_tracking(db)
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Relatório de Rastreamento", ln=True, align='C')

    for track in tracking_data:
        pdf.cell(200, 10, txt=f"ID: {track.id}, Material ID: {track.material_id}, Etapa: {track.stage}, Status: {track.status}, Data: {track.created_at}", ln=True)

    file_path = "/tmp/report.pdf"
    pdf.output(file_path)
    return file_path

# Gera um relatório XLSX de rastreamento
def generate_xlsx_report(db: Session):
    tracking_data = get_tracking(db)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Rastreamento"

    headers = ["ID", "Material ID", "Etapa", "Status", "Data"]
    sheet.append(headers)

    for track in tracking_data:
        sheet.append([track.id, track.material_id, track.stage, track.status, track.created_at])

    file_path = "/tmp/report.xlsx"
    workbook.save(file_path)
    return file_path